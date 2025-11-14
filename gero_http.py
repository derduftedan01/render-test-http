import os
import pandas as pd


def process_data(input_data: dict):
    return {
        "received": input_data,
        "message": "Gero hat die Daten verarbeitet."
    }

def geronimo_excel():
    pd.set_option("display.max_columns", None)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_dir, "geronimo_input.xlsx")

    days = pd.read_excel(excel_path, sheet_name="arbeitstage")
    data = pd.read_excel(excel_path, sheet_name="rohdaten")
    dash = pd.DataFrame(columns=["Festnetz", "Arbeitstage", "Gesamtcalls", "Gesamtcalls pro AT", "Nettocalls gesamt",
                                "Outbound Nettocalls", "Outbound gesamt", "Outbound Callduration", "Outbound Ɵ Callduration", "Outbound intern", "Outbound extern", "Outbound Unanswered",
                                "Inbound Nettocalls", "Inbound gesamt", "Inbound Callduration", "Inbound Ɵ Callduration", "Inbound intern", "Inbound extern", "Inbound Unanswered",
                                "Nettocalls pro AT", "Netto Outbound pro AT", "Netto Inbound pro AT", "Callduration gesamt", "Ɵ Callduration gesamt", "Verhältnis Outbound", "Unanswered Outbound", "Unanswered Inbound",
                                "Outbound < 10 sek", "Outbound < 1 min", "Outbound 1-2 min", "Outbound 2-3 min", "Outbound > 3 min", "Outbound % > 3 min", "Outbound > 3 min pro AT", "Inbound < 10 sek", "Inbound < 1 min", "Inbound 1-2 min", "Inbound 2-3 min", "Inbound > 3 min", "Inbound % > 3 min", "Inbound > 3 min pro AT"
    ])

    data.rename(columns={"Start ": "Start"}, inplace=True)
    data["Start"] = pd.to_datetime(data["Start"]).dt.date
    data["Calling Number"] = data["Calling Number"].astype(str)
    data["Called Number"] = data["Called Number"].astype(str)
    data["Duration (s)"] = pd.to_numeric(data["Duration (s)"], errors="coerce").fillna(0)
    days["Arbeitstage"] = pd.to_numeric(days["Arbeitstage"], errors="coerce").fillna(0)
    days["Festnetz"] = days["Festnetz"].astype(str)

    def safe_div(a, b):
        import numpy as np

        a_arr = np.asarray(a, dtype="float64")
        b_arr = np.asarray(b, dtype="float64")

        b_arr = np.where(b_arr == 0, np.nan, b_arr)

        with np.errstate(divide="ignore", invalid="ignore"):
            res = a_arr / b_arr

        res[~np.isfinite(res)] = 0

        if isinstance(a, pd.Series):
            return pd.Series(res, index=a.index)
        return res

    dash = pd.concat([dash, days], ignore_index=True)
    for _, drow in dash.iterrows():

        arbeitstage = max(drow["Arbeitstage"], 1)

        festnetz = drow["Festnetz"]
        if festnetz not in data["Calling Number"].values and festnetz not in data["Called Number"].values:
            continue

        dash.loc[dash["Festnetz"] == festnetz, "Outbound gesamt"] =         len(data[data["Calling Number"] == festnetz])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound intern"] =         len(data[(data["Calling Number"] == festnetz) & (data["Called Number"].astype(str).str.startswith("4922347011"))])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound extern"] =         len(data[(data["Calling Number"] == festnetz) & (~data["Called Number"].astype(str).str.startswith("4922347011"))])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound Unanswered"] =     len(data[(data["Calling Number"] == festnetz) & (data["Answered"] == "Unanswered")])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound Nettocalls"] =     dash.loc[dash["Festnetz"] == festnetz, "Outbound extern"] - dash.loc[dash["Festnetz"] == festnetz, "Outbound Unanswered"]
        dash.loc[dash["Festnetz"] == festnetz, "Outbound Callduration"] =   (data[(data["Calling Number"] == festnetz) & (~data["Called Number"].astype(str).str.startswith("4922347011"))]["Duration (s)"].sum()) / 60
        dash.loc[dash["Festnetz"] == festnetz, "Outbound Ɵ Callduration"] = safe_div(dash.loc[dash["Festnetz"] == festnetz, "Outbound Callduration"], dash.loc[dash["Festnetz"] == festnetz, "Outbound Nettocalls"])

        dash.loc[dash["Festnetz"] == festnetz, "Inbound gesamt"] =          len(data[data["Called Number"] == festnetz])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound intern"] =          len(data[(data["Called Number"] == festnetz) & (data["Calling Number"].astype(str).str.startswith("4922347011"))])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound extern"] =          len(data[(data["Called Number"] == festnetz) & (~data["Calling Number"].astype(str).str.startswith("4922347011"))])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound Unanswered"] =      len(data[(data["Called Number"] == festnetz) & (data["Answered"] == "Unanswered")])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound Nettocalls"] =      dash.loc[dash["Festnetz"] == festnetz, "Inbound extern"] - dash.loc[dash["Festnetz"] == festnetz, "Inbound Unanswered"]
        dash.loc[dash["Festnetz"] == festnetz, "Inbound Callduration"] =    (data[(data["Called Number"] == festnetz) & (~data["Calling Number"].astype(str).str.startswith("4922347011"))]["Duration (s)"].sum()) / 60
        dash.loc[dash["Festnetz"] == festnetz, "Inbound Ɵ Callduration"] =  safe_div(dash.loc[dash["Festnetz"] == festnetz, "Inbound Callduration"], dash.loc[dash["Festnetz"] == festnetz, "Inbound Nettocalls"])

        dash.loc[dash["Festnetz"] == festnetz, "Gesamtcalls"] =             dash.loc[dash["Festnetz"] == festnetz, "Outbound gesamt"] + dash.loc[dash["Festnetz"] == festnetz, "Inbound gesamt"]
        dash.loc[dash["Festnetz"] == festnetz, "Gesamtcalls pro AT"] =      safe_div(dash.loc[dash["Festnetz"] == festnetz, "Gesamtcalls"], arbeitstage)
        dash.loc[dash["Festnetz"] == festnetz, "Nettocalls gesamt"] =       dash.loc[dash["Festnetz"] == festnetz, "Outbound Nettocalls"] + dash.loc[dash["Festnetz"] == festnetz, "Inbound Nettocalls"]

        print("arbeitstage: ", arbeitstage)
        print("Nettocalls gesamt: ", dash.loc[dash["Festnetz"] == festnetz, "Nettocalls gesamt"])
        print("Outbound Nettocalls: ", dash.loc[dash["Festnetz"] == festnetz, "Outbound Nettocalls"])
        print("Inbound Nettocalls: ", dash.loc[dash["Festnetz"] == festnetz, "Inbound Nettocalls"])
        dash.loc[dash["Festnetz"] == festnetz, "Nettocalls pro AT"] =       safe_div(dash.loc[dash["Festnetz"] == festnetz, "Nettocalls gesamt"], arbeitstage)
        dash.loc[dash["Festnetz"] == festnetz, "Netto Outbound pro AT"] =   safe_div(dash.loc[dash["Festnetz"] == festnetz, "Outbound Nettocalls"], arbeitstage)
        dash.loc[dash["Festnetz"] == festnetz, "Netto Inbound pro AT"] =    safe_div(dash.loc[dash["Festnetz"] == festnetz, "Inbound Nettocalls"], arbeitstage)
        dash.loc[dash["Festnetz"] == festnetz, "Callduration gesamt"] =     dash.loc[dash["Festnetz"] == drow["Festnetz"], "Outbound Callduration"] + dash.loc[dash["Festnetz"] == drow["Festnetz"], "Inbound Callduration"]
        dash.loc[dash["Festnetz"] == festnetz, "Ɵ Callduration gesamt"] =   safe_div(dash.loc[dash["Festnetz"] == festnetz, "Callduration gesamt"], dash.loc[dash["Festnetz"] == festnetz, "Nettocalls gesamt"])
        dash.loc[dash["Festnetz"] == festnetz, "Verhältnis Outbound"] =     safe_div(dash.loc[dash["Festnetz"] == festnetz, "Outbound Nettocalls"], dash.loc[dash["Festnetz"] == festnetz, "Nettocalls gesamt"])
        dash.loc[dash["Festnetz"] == festnetz, "Unanswered Outbound"] =     safe_div(dash.loc[dash["Festnetz"] == festnetz, "Outbound Unanswered"], dash.loc[dash["Festnetz"] == festnetz, "Outbound gesamt"])
        dash.loc[dash["Festnetz"] == festnetz, "Unanswered Inbound"] =      safe_div(dash.loc[dash["Festnetz"] == festnetz, "Inbound Unanswered"], dash.loc[dash["Festnetz"] == festnetz, "Inbound gesamt"])

        dash.loc[dash["Festnetz"] == festnetz, "Outbound < 10 sek"] =       len(data[(data["Calling Number"] == festnetz) & (~data["Called Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] > 0) & (data["Duration (s)"] < 10)])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound < 1 min"] =        len(data[(data["Calling Number"] == festnetz) & (~data["Called Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] >= 10) & (data["Duration (s)"] < 60)])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound 1-2 min"] =        len(data[(data["Calling Number"] == festnetz) & (~data["Called Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] >= 60) & (data["Duration (s)"] < 120)])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound 2-3 min"] =        len(data[(data["Calling Number"] == festnetz) & (~data["Called Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] >= 120) & (data["Duration (s)"] < 180)])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound > 3 min"] =        len(data[(data["Calling Number"] == festnetz) & (~data["Called Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] >= 180)])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound % > 3 min"] =      safe_div(dash.loc[dash["Festnetz"] == festnetz, "Outbound > 3 min"], dash.loc[dash["Festnetz"] == festnetz, "Outbound Nettocalls"])
        dash.loc[dash["Festnetz"] == festnetz, "Outbound > 3 min pro AT"] = safe_div(dash.loc[dash["Festnetz"] == festnetz, "Outbound > 3 min"], arbeitstage)

        dash.loc[dash["Festnetz"] == festnetz, "Inbound < 10 sek"] =        len(data[(data["Called Number"] == festnetz) & (~data["Calling Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] > 0) & (data["Duration (s)"] < 10)])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound < 1 min"] =         len(data[(data["Called Number"] == festnetz) & (~data["Calling Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] >= 10) & (data["Duration (s)"] < 60)])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound 1-2 min"] =         len(data[(data["Called Number"] == festnetz) & (~data["Calling Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] >= 60) & (data["Duration (s)"] < 120)])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound 2-3 min"] =         len(data[(data["Called Number"] == festnetz) & (~data["Calling Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] >= 120) & (data["Duration (s)"] < 180)])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound > 3 min"] =         len(data[(data["Called Number"] == festnetz) & (~data["Calling Number"].astype(str).str.startswith("4922347011")) & (data["Duration (s)"] >= 180)])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound % > 3 min"] =       safe_div(dash.loc[dash["Festnetz"] == festnetz, "Inbound > 3 min"], dash.loc[dash["Festnetz"] == festnetz, "Inbound Nettocalls"])
        dash.loc[dash["Festnetz"] == festnetz, "Inbound > 3 min pro AT"] =  safe_div(dash.loc[dash["Festnetz"] == festnetz, "Inbound > 3 min"], arbeitstage)

    num_cols = dash.columns.drop("Festnetz")
    dash[num_cols] = dash[num_cols].apply(pd.to_numeric, errors='coerce')

    sum_row = dash[num_cols].sum()
    sum_row["Festnetz"] = "Σ Summe"

    mean_row = dash[num_cols].mean()
    mean_row["Festnetz"] = "Ø Mittelwert"

    dash = pd.concat([dash, pd.DataFrame([sum_row]), pd.DataFrame([mean_row])], ignore_index=True)



    # daily calls per number
    out = data.loc[data["Direction"].eq("Outbound")].copy()

    out["Datum"] = pd.to_datetime(out["Start"]).dt.date

    nums = pd.Index(days["Festnetz"]).astype(str).unique()
    out["Calling Number"] = out["Calling Number"].astype(str)

    counts = (
        out.groupby(["Datum", "Calling Number"])
        .size()
        .unstack("Calling Number", fill_value=0)
        .reindex(columns=nums, fill_value=0)
    )

    #print("counts: ", counts)

    all_days = pd.to_datetime(data["Start"]).dt.date
    counts = counts.reindex(sorted(pd.unique(all_days)), fill_value=0)
    counts.index.name = "Datum"

    #print("counts: ", counts)

    daily = (
        counts
        .reset_index()
        .sort_values("Datum")
        .reset_index(drop=True)
    )

    #print("daily: ", daily)

    daily["Wochentag"] = pd.to_datetime(daily["Datum"]).dt.day_name(locale="de_DE")

    value_cols = [c for c in daily.columns if c not in ("Datum", "Wochentag")]
    daily = daily[["Datum", "Wochentag"] + value_cols]




    out_netto = data.loc[data["Direction"].eq("Outbound") & data["Answered"].eq("Answered")].copy()

    out_netto["Datum"] = pd.to_datetime(out_netto["Start"]).dt.date

    nums = pd.Index(days["Festnetz"]).astype(str).unique()
    out_netto["Calling Number"] = out_netto["Calling Number"].astype(str)

    counts_netto = (
        out_netto.groupby(["Datum", "Calling Number"])
        .size()
        .unstack("Calling Number", fill_value=0)
        .reindex(columns=nums, fill_value=0)
    )

    #print("counts_netto: ", counts_netto)

    all_days = pd.to_datetime(data["Start"]).dt.date
    counts_netto = counts_netto.reindex(sorted(pd.unique(all_days)), fill_value=0)
    counts_netto.index.name = "Datum"

    #print("counts_netto: ", counts_netto)

    daily_netto = (
        counts_netto
        .reset_index()
        .sort_values("Datum")
        .reset_index(drop=True)
    )

    #print("daily_netto: ", daily_netto)

    daily_netto["Wochentag"] = pd.to_datetime(daily_netto["Datum"]).dt.day_name(locale="de_DE")

    value_cols_netto = [c for c in daily_netto.columns if c not in ("Datum", "Wochentag")]
    daily_netto = daily_netto[["Datum", "Wochentag"] + value_cols_netto]



    # Spaltenliste der Nummern (ohne Datum/Wochentag)
    value_cols = [c for c in daily.columns if c not in ("Datum", "Wochentag")]

    # Auf Datum ausrichten und gleiche Spaltenreihenfolge sicherstellen
    df_b = daily.sort_values("Datum").set_index("Datum")[value_cols]
    df_n = daily_netto.sort_values("Datum").set_index("Datum")[value_cols]

    # Optional: feste Reihenfolge gemäss nums
    df_b = df_b.reindex(columns=[str(c) for c in nums], fill_value=0)
    df_n = df_n.reindex(columns=[str(c) for c in nums], fill_value=0)

    # Zusammenfügen: ('Brutto'/'Netto') x Nummern
    wide = pd.concat({"Brutto": df_b, "Netto": df_n}, axis=1)
    # Ebenen tauschen -> oben Nummer, unten Brutto/Netto
    wide = wide.swaplevel(0, 1, axis=1).sort_index(axis=1, level=0)

    # Datum/Wochentag wieder hinzufügen
    wide = wide.reset_index()
    wide["Wochentag"] = pd.to_datetime(wide["Datum"]).dt.day_name(locale="de_DE")

    # Einzeilige Köpfe für Datum/Wochentag + MultiIndex für Nummern
    #first = [("Datum", ""), ("Wochentag", "")]
    #rest  = [(str(a), b) for (a, b) in wide.columns if a not in ("Datum", "Wochentag")]

    wide.columns = pd.MultiIndex.from_tuples(
        [c if isinstance(c, tuple) else (str(c), "") for c in wide.columns]
    )

    wt = wide.pop("Wochentag")
    wide.insert(1, "Wochentag", wt)

    fixed = [("Datum", ""), ("Wochentag", "")]
    rest  = [c for c in wide.columns if c not in fixed]
    wide = wide.loc[:, fixed + rest]

    wide = wide.sort_index(axis=1, level=0, ascending=False)

    output_path = os.path.join(base_dir, "geronimo_output.xlsx")

    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        data.to_excel(writer, sheet_name="rohdaten", index=False)
        days.to_excel(writer, sheet_name="arbeitstage", index=False)
        dash.to_excel(writer, sheet_name="auswertung", index=False)
        wide.to_excel(writer, sheet_name="tagescalls", index=True, merge_cells=True)



        ws = writer.sheets["tagescalls"]
        ws.set_column(0, 0, None, None, {'hidden': True})

        all_ws = writer.sheets

        for sheet_name, ws in all_ws.items():
            for i, col in enumerate(daily.columns):
                max_len = len(str(col))
                col_len = daily[col].astype(str).map(len).max()
                if col_len > max_len:
                    max_len = col_len
                ws.set_column(i, i, max_len + 2)



    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string

    wb = load_workbook(output_path)
    ws = wb["auswertung"]

    # format cells
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    cell_alignment = Alignment(
        horizontal='center',      # Horizontal zentrieren
        vertical='center'         # Vertikal zentrieren
    )

    for row in ws.iter_rows(min_row=2, max_row=9, min_col=1, max_col=ws.max_column):
        for i, cell in enumerate(row, start=1):
            cell.alignment = cell_alignment
            cell.border = thin_border
            if i in [1, 2, 3, 5, 6, 7, 10, 11, 12, 13, 14, 17, 18, 19, 28, 29, 30, 31, 32, 35, 36, 37, 38, 39]:
                cell.number_format = '0'
            elif i in [25, 26, 27, 33, 40]:
                cell.number_format = "0.00%"
            elif i in [8, 9, 15, 16, 23, 24]:
                minutes = float(cell.value) if cell.value is not None else 0
                excel_time = minutes / (24 * 60)
                cell.value = excel_time
                cell.number_format = "hh:mm:ss" #[mm]:ss
            else:
                cell.number_format = '0.00'

    # row height and alignment
    ws.row_dimensions[1].height = 45
    header_alignment = Alignment(
        wrap_text=True,           # Zeilenumbruch aktivieren
        horizontal='center',      # Horizontal zentrieren
        vertical='center'         # Vertikal zentrieren
    )

    for cell in ws[1]:
        cell.alignment = header_alignment

    # column widths
    max_columns = ws.max_column
    for i in range(1, max_columns + 1):
        col_letter = get_column_letter(i)

        if i == 1:
            width = 16
        elif i >= 28:
            width = 10
        else:
            width = 12

        ws.column_dimensions[col_letter].width = width

    # colouring
    color_groups = {
        "single_blue": {
            "columns": [1],  # A
            "header": "002060",
            "data": None
        },
        "dark_green": {
            "columns": list(range(2, 6)),  # B-D
            "header": "385723",
            "data": "EBF1DE"
        },
        "dark_blue": {
            "columns": list(range(6, 13)),  # E-L
            "header": "002060",
            "data": "DAEEF3"
        },
        "light_blue": {
            "columns": list(range(13, 20)),  # M-S
            "header": "1F497D",
            "data": "DCE6F1"
        },
        "purple": {
            "columns": list(range(20, 28)),  # T–AA
            "header": "7030A0",
            "data": "E4DFEC"
        },
        "magenta": {
            "columns": list(range(28, 42)),  # AB-AO
            "header": "A10054",
            "data": "FDE9F1"
        },
    }

    font_white = Font(color="FFFFFF", bold=True)

    def apply_group_coloring(ws):
        special_rows = [1, 8, 9]
        data_rows = range(2, 8)

        for group in color_groups.values():
            # Farben vorbereiten
            header_fill = PatternFill(start_color=group["header"], end_color=group["header"], fill_type="solid")
            data_fill = None
            if group.get("data"):
                data_fill = PatternFill(start_color=group["data"], end_color=group["data"], fill_type="solid")

            for col in group["columns"]:
                # Header + Summe + Ø
                for row in special_rows:
                    cell = ws.cell(row=row, column=col)
                    cell.fill = header_fill
                    cell.font = font_white
                    cell.border = thin_border

                # Datenzeilen
                if data_fill:
                    for row in data_rows:
                        ws.cell(row=row, column=col).fill = data_fill
                        cell.border = thin_border

    apply_group_coloring(ws)

    # black columns
    insert_cols_original = [2, 6, 13, 20, 28]

    for col_index in sorted(insert_cols_original, reverse=True):
        ws.insert_cols(col_index)

    black_cols = [
        orig_index + i
        for i, orig_index in enumerate(sorted(insert_cols_original))
    ]

    black_header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    black_header_font = Font(color="FFFFFF", bold=True)

    for col in black_cols:
        ws.column_dimensions[get_column_letter(col)].width = 6
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = black_header_fill
            cell.font = black_header_font
            cell.border = thin_border

    # export
    wb.save(output_path)